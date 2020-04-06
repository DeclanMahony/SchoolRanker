# -*- coding: utf-8 -*-
"""
Created on Tue Mar 31 07:26:16 2020

@author: decla
"""


import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import openpyxl
import xlwt
import numpy
import math
from sklearn.linear_model import LinearRegression
from sympy import symbols, solve, diff
import pandas

# Load the Excel file and access the Maths and Language results sheets.
#file = input('Enter the file directory')
file = "ADvTECH Gr8 2020 (Dex).xlsx"
workbook = openpyxl.load_workbook(file, data_only=True)

# Input the raw data from a given sheet
def input_data(sheet_name, headings_row):

	sheet = workbook[sheet_name]
	
# Input the raw data
	# Fetch the headings
	next_cell = ''
	headings = []
	count = 1
	while next_cell != None:
		headings.append(sheet.cell(row=headings_row,column=count).value)
		count+=1
		next_cell = sheet.cell(row=headings_row,column=count).value
	
	# Read in the structured data in this format:
	'''
	data = {School 1:{
			Grade 1:{
				Class 1:{
					Teacher:___, 
					Students:{
						Student 1:{ 
							First Name:___, 
							Surname:___, 
							S/No:___,
							Oldest:___, 
							Most Recent:___, 
							Scores:{
									Q1:{
										Mark:___,
										Grade Level:___,
										Cognitive Domain:___,
										Content Domain:___}
									Q2:{
									.
									.
									.
									}}}
						Student 2:{
							.
							.
							.}}}
				Class 2:{
					.
					.
					.
					}}
			Grade 2:{
				.
				.
				.}}
			School 2:{
				.
				.
				.
				}}
	'''

	data={}
	next_school=''
	current_cell=''
	count=headings_row+1
	# read in all the school names
	while next_school != None:
		current_cell = sheet.cell(row=count, column = 1).value
		if current_cell not in data:
			data[current_cell]={}
		count+=1
		next_school = sheet.cell(row=count, column = 1).value

	next_school=''
	current_cell=''
	count=headings_row+1
	# read in all the grades
	for school in data:
		next_school = sheet.cell(row=count+1, column = 1).value
		while next_school == school:
			current_cell = sheet.cell(row=count, column = 2).value
			if current_cell not in data[school]:
				data[school][current_cell]={}
			count+=1
			next_school = sheet.cell(row=count, column = 1).value
			
	# read in all the classes
	next_school=''
	next_grade=''
	current_cell=''
	count = headings_row+1
	for school in data:
		for grade in data[school]:
			next_school = sheet.cell(row=count+1, column = 1).value
			next_grade = sheet.cell(row=count+1, column = 2).value
			while next_grade == grade and next_school == school:
				current_cell = sheet.cell(row=count, column = 3).value
				if current_cell not in data[school][grade]:
					data[school][grade][current_cell]={}
				count+=1
				next_grade=sheet.cell(row=count, column = 2).value
				next_school = sheet.cell(row=count, column = 1).value
				
	# read in all the teachers and students
	count = headings_row+1
	for school in data:
		for grade in data[school]:
			for Class in data[school][grade]:
				next_class = sheet.cell(row=count+1, column = 3).value
				next_school = sheet.cell(row=count+1, column = 1).value
				next_grade = sheet.cell(row=count+1, column = 2).value
				while next_grade == grade and next_school == school and next_class == Class:
					data[school][grade][Class].update({'Teacher': {'Name':sheet.cell(row=count, column = 4).value},\
		  sheet.cell(row=count, column = 8).value:{\
											  'School': sheet.cell(row=count, column = 1).value,\
											  'Grade':sheet.cell(row=count, column = 2).value,\
											  'Class':sheet.cell(row=count, column = 3).value,\
											  'First name': sheet.cell(row=count, column = 6).value,\
											  'Surname': sheet.cell(row=count, column = 7).value,\
											  'Language': sheet.cell(row=count, column = 5).value,\
											  'Oldest': sheet.cell(row=count, column = 9).value,\
											  'Most recent': sheet.cell(row=count, column = 10).value,\
											  'Device': sheet.cell(row=count, column = 11).value,\
											  'Scores':{},\
											  'Number':count}})
					count+=1
					next_class = sheet.cell(row=count, column = 3).value
					next_grade=sheet.cell(row=count, column = 2).value
					next_school = sheet.cell(row=count, column = 1).value				
	
	# read in the students' scores
	
	for school in data:
		for grade in data[school]:
			for Class in data[school][grade]:
				for student in data[school][grade][Class]:
					if student != 'Teacher':
						count = headings.index('Q1')+1
						next_cell = sheet.cell(row = headings_row , column = count).value
						while next_cell[0] == 'Q' and next_cell[1].isnumeric():
							if sheet.cell(row = data[school][grade][Class][student]['Number'], column = count).value == '-':
								data[school][grade][Class][student]['Scores'].update({sheet.cell(row = headings_row , column = count).value:{\
									     'Mark':0,\
										 'Grade Level': sheet.cell(row = headings_row -1 , column = count).value,\
										 'Cognitive Domain': sheet.cell(row = headings_row -2 , column = count).value,\
										 'Content Domain':sheet.cell(row = headings_row -3 , column = count).value}})
							else:
								data[school][grade][Class][student]['Scores'].update({sheet.cell(row = headings_row , column = count).value:{\
										     'Mark':sheet.cell(row = data[school][grade][Class][student]['Number'] , column = count).value,\
											 'Grade Level': sheet.cell(row = headings_row -1 , column = count).value,\
											 'Cognitive Domain': sheet.cell(row = headings_row -2 , column = count).value,\
											 'Content Domain':sheet.cell(row = headings_row -3 , column = count).value}})
							count+=1
							next_cell = sheet.cell(row = headings_row , column = count).value
				
	return data

# Rank each student according to grade/cognitive/content
def rank_students(data,rank,threshold):
	ranks=[]
	student_ranks = {}
	student_scores={}
	student_averages={}
	for school in data:
		student_scores[school]={}
		for grade in data[school]:
			for Class in data[school][grade]:
				for student in data[school][grade][Class]:
					if student != 'Teacher':
						student_scores[school][student]={'Details':data[school][grade][Class][student]}
						for question in data[school][grade][Class][student]['Scores']:
							if data[school][grade][Class][student]['Scores'][question][rank] not in student_scores[school][student]:
								student_scores[school][student].update({data[school][grade][Class][student]['Scores'][question][rank]:[data[school][grade][Class][student]['Scores'][question]['Mark']]})
							else:
								student_scores[school][student][data[school][grade][Class][student]['Scores'][question][rank]].append(int(data[school][grade][Class][student]['Scores'][question]['Mark']))
	for school in student_scores:
		student_averages[school]={}
		for student in student_scores[school]:
			if student != 'Teacher':
				student_averages[school][student]={'Details':student_scores[school][student]['Details']}
				for rank in student_scores[school][student]:
					if rank!='Details':
						if rank not in ranks:
							ranks.append(rank)
						student_averages[school][student][rank]=sum(student_scores[school][student][rank])/len(student_scores[school][student][rank])*100					
	for school in student_averages:
		student_ranks[school]={}
		for student in student_averages[school]:
			student_ranks[school][student]={'Details':student_averages[school][student]['Details']}
			for i in ranks[::-1]:
				if student_averages[school][student][i]>=threshold and i != ranks[0]:
					student_ranks[school][student]['Rank']=i
					break
				elif i == ranks[0]:
					student_ranks[school][student]['Rank']=i
				
	return student_ranks, student_averages, student_scores, ranks
''' A dictionary of the form student_ranks = {school 1:{\
																				student 1:{\
																						   rank:___},
																				student 2:{\
																							rank:___},
																					.
																					.
																					.}
																		school 2:{
																			.
																			.
																			.}}
'''
# rank the schools
def rank_schools(student_ranks,ranks,grade):
	school_ranks={}
	for school in student_ranks:
		school_ranks[school]={}
		for rank in ranks:
			school_ranks[school][rank]=0
		for student in student_ranks[school]:
			school_ranks[school][student_ranks[school][student]['Rank']]+=1
		school_ranks[school]['Number of students']=len(student_ranks[school])
	for school in school_ranks:
		for rank in school_ranks[school]:
			school_ranks[school][rank]=school_ranks[school][rank]/school_ranks[school]['Number of students']*100
		school_ranks[school]['Grade Rank']=school_ranks[school]['G'+str(grade)]+school_ranks[school]['G'+str(grade-1)]
		school_ranks[school]['Number of students']=len(student_ranks[school])
	return school_ranks
maths_data=input_data('M8',11)
language_data=input_data('L8',10)		

#Learner ranks according to Grade Level
threshold = int(input("What's the threshold?"))
maths_ranks, maths_averages, maths_scores, maths_ticks =rank_students(maths_data,'Grade Level',threshold)
language_ranks, language_averages, language_scores, language_ticks = rank_students(language_data, 'Grade Level',threshold)		

school_rank_maths=rank_schools(maths_ranks, maths_ticks,8)
school_rank_language=rank_schools(language_ranks,language_ticks,8)

# lets plot some graphs
maths_x = numpy.arange(len(school_rank_maths))
language_x = numpy.arange(len(school_rank_language))

maths_y_prev=[0]*len(maths_x)
for rank in maths_ticks:
	maths_y=[]
	for school in school_rank_maths:
		if rank in school_rank_maths[school]:
			maths_y.append(school_rank_maths[school][rank])
		else:
			maths_y.append(0)
	plt.bar(maths_x,maths_y,bottom=maths_y_prev)
	maths_y_prev = numpy.add(maths_y,maths_y_prev)
plt.show()

language_y_prev=[0]*len(language_x)
for rank in language_ticks:
	language_y=[]
	for school in school_rank_language:
		if rank in school_rank_language[school]:
			language_y.append(school_rank_language[school][rank])
		else:
			language_y.append(0)
	plt.bar(language_x,language_y,bottom=language_y_prev)
	language_y_prev = numpy.add(language_y,language_y_prev)
plt.show()

# Combine language and maths ranks
formatted={}
for school1 in school_rank_maths:
	for school2 in school_rank_language:
		if school1 == school2:
			formatted[school1]={}
			formatted[school1]['Rank']=school_rank_maths[school1]['Grade Rank'] + school_rank_language[school1]['Grade Rank']
			for rank in school_rank_maths[school1]:
				formatted[school1]['M'+str(rank)]=school_rank_maths[school1][rank]
			for rank in school_rank_language[school1]:
				formatted[school1]['L'+str(rank)]=school_rank_language[school1][rank]

		elif school1 not in school_rank_language:
			formatted[school1]={}
			formatted[school1]['Rank']=school_rank_maths[school1]['Grade Rank']
			for rank in school_rank_maths[school1]:
				formatted[school1]['M'+str(rank)]=school_rank_maths[school1][rank]
			for rank in school_rank_language[school2]:
				formatted[school1]['L'+str(rank)]=None

		elif school2 not in school_rank_maths:
			formatted[school2]={}
			formatted[school2]['Rank']=school_rank_language[school2]['Grade Rank']
			for rank in school_rank_maths[school1]:
				formatted[school2]['M'+str(rank)]=0
			for rank in school_rank_language[school2]:
				formatted[school2]['L'+str(rank)]=school_rank_language[school2][rank]


#Write data to excel using openpyxl

new_book = openpyxl.Workbook()
sheet1=new_book.active
sheet1.title = 'School Rank'
sheet1.cell(row=1,column=1).value =  'School'
for r,school in enumerate(formatted):
	sheet1.cell(row=r+2,column = 1).value = school
	for c, rank in enumerate(formatted[school]):
		sheet1.cell(row = 1, column = c +2).value = rank
		sheet1.cell(row = r +2, column = c+2).value = formatted[school][rank]
	


# Bubble student averages to the school level
school_averages={}
for school1 in maths_averages:
	for school2 in language_averages:
		if school1==school2:
			school_averages[school1]={}
			for student in maths_averages[school1]:
				for grade in maths_averages[school1][student]:
					if grade!='Details':
						if 'M'+grade not in school_averages[school1]:
							school_averages[school1]['M'+grade]=maths_averages[school1][student][grade]/school_rank_maths[school1]['Number of students']
						else:
							school_averages[school1]['M'+grade]+=maths_averages[school1][student][grade]/school_rank_maths[school1]['Number of students']
			for student in language_averages[school1]:	
				for grade in language_averages[school1][student]:
					if grade !='Details':
						if 'L'+grade not in school_averages[school1]:
							school_averages[school1]['L'+grade]=language_averages[school1][student][grade]/school_rank_language[school1]['Number of students']
						else:
							school_averages[school1]['L'+grade]+=language_averages[school1][student][grade]/school_rank_language[school1]['Number of students']
			
		elif school1 not in language_averages:
			school_averages[school1]={}
			for student in maths_averages[school1]:
				for grade in maths_averages[school1][student]:
					if grade != 'Details':
						if 'M'+grade not in school_averages[school1]:
							school_averages[school1]['M'+grade]=maths_averages[school1][student][grade]/school_rank_maths[school1]['Number of students']
						else:
							school_averages[school1]['M'+grade]+=maths_averages[school1][student][grade]/school_rank_maths[school1]['Number of students']
			for learner in language_averages[school2]:
				for grade in language_averages[school2][learner]:
					if grade!='Details':
						school_averages[school1]['L'+grade]=0
					
		elif school2 not in maths_averages:
			school_averages[school2]={}
			for student in language_averages[school2]:
				for grade in language_averages[school2][student]:
					if grade!='Details':
						if 'L'+grade not in school_averages[school1]:
							school_averages[school2]['L'+grade]=maths_averages[school2][student][grade]/school_rank_language[school2]['Number of students']
						else:
							school_averages[school2]['L'+grade]+=maths_averages[school2][student][grade]/school_rank_language[school2]['Number of students']
			for learner in maths_averages[school1]:
				for grade in maths_averages[school1][learner]:
					if grade!='Details':
						school_averages[school2]['M'+grade]=0		
	
sheet2=new_book.create_sheet('School Averages')
sheet2.cell(row=1,column=1).value= 'School'
for r,school in enumerate(school_averages):
	sheet2.cell(row=r+2,column=1).value=school
	for c,grade in 	enumerate(school_averages[school]):
		sheet2.cell(row=1,column=c+2).value=grade
		sheet2.cell(row=r+2,column=c+2).value=school_averages[school][grade]			

# Learner ranks according to cognitive and content domains
cognitive_maths_ranks, cognitive_maths_averages, cognitive_maths_scores, cognitive_maths_ticks =rank_students(maths_data,'Cognitive Domain',threshold)
cognitive_language_ranks, cognitive_language_averages, cognitive_language_scores, cognitive_language_ticks = rank_students(language_data, 'Cognitive Domain', threshold)	

content_maths_ranks, content_maths_averages, content_maths_scores, content_maths_ticks =rank_students(maths_data,'Content Domain',threshold)
content_language_ranks, content_language_averages, content_language_scores, content_language_ticks = rank_students(language_data, 'Content Domain', threshold)

# Combine maths and language averages for grades and cognitive levels for each learner
student_averages={}
for school1 in maths_averages:
	for school2 in language_averages:
		for student1 in maths_averages[school1]:
			for student2 in language_averages[school2]:
				if student1==student2:
					
					student_averages[student1]={'Details':maths_averages[school1][student1]['Details'],\
												  'Maths':{'Grade Levels': maths_averages[school1][student1],\
														    'Cognitive Domain':cognitive_maths_averages[school1][student1],\
															'Content Domain':content_maths_averages[school1][student1],\
															'Average':{}},\
												   'Language':{'Grade Levels': language_averages[school2][student2],\
															   'Cognitive Domain': cognitive_language_averages[school2][student2],\
															   'Content Domain': content_language_averages[school2][student2],\
															   'Average':{}},\
												   'Overall Average':{}}
				     

				elif student1 not in student_averages and student1!=student2 and student1 not in language_averages:
					
					student_averages[student1]={'Details':maths_averages[school1][student1]['Details'],\
												  'Maths':{'Grade Levels': maths_averages[school1][student1],\
														    'Cognitive Domain':cognitive_maths_averages[school1][student1],\
															'Content Domain':content_maths_averages[school1][student1],\
															'Average':{}},\
												   'Language':{'Grade Levels': {},\
															   'Cognitive Domain': {},\
															   'Content Domain': {},\
															   'Average':{},\
												   'Overall Average':{}}}

				elif student2 not in student_averages and student1!=student2 and student2 not in maths_averages:
					
					student_averages[student2]={'Details':language_averages[school2][student2]['Details'],\
												  'Maths':{'Grade Levels':{},\
														    'Cognitive Domain':{},\
															'Content Domain':{},\
															'Average':{}},\
												   'Language':{'Grade Levels': language_averages[school2][student2],\
															   'Cognitive Domain': cognitive_language_averages[school2][student2],\
															   'Content Domain': content_language_averages[school2][student2],\
															   'Average':{},\
												   'Overall Average':{}}}
# Calculate Averages
for student in student_averages:
	m_average = 0
	for grade in student_averages[student]['Maths']['Grade Levels']:
		if grade!='Details':
			m_average+=student_averages[student]['Maths']['Grade Levels'][grade]/(len(student_averages[student]['Maths']['Grade Levels'])-1)
	student_averages[student]['Maths']['Average']=m_average
	l_average = 0
	for grade in student_averages[student]['Language']['Grade Levels']:
		if grade!='Details':
			l_average+=student_averages[student]['Language']['Grade Levels'][grade]/(len(student_averages[student]['Language']['Grade Levels'])-1)
	student_averages[student]['Language']['Average']=l_average 
	student_averages[student]['Overall Average'] = (l_average+m_average)/2

# Write student average data to excel
sheet3=new_book.create_sheet('Student Averages')
sheet4=new_book.create_sheet('Grade Level Averages')
sheet5=new_book.create_sheet('Cognitive Domain Averages')
sheet6=new_book.create_sheet('Content Domain Averages')

sheet3.cell(row=2,column = 2+len(student_averages[student]['Details'])).value = 'Maths Average'
sheet3.cell(row=2,column = 3+len(student_averages[student]['Details'])).value = 'Language Average'
sheet3.cell(row=2,column = 4+len(student_averages[student]['Details'])).value = 'Overall Average'

for r,student in enumerate(student_averages):
	
	sheet3.cell(row=r+3, column=2+len(student_averages[student]['Details'])).value = student_averages[student]['Maths']['Average']
	sheet3.cell(row=r+3, column=3+len(student_averages[student]['Details'])).value = student_averages[student]['Language']['Average']
	sheet3.cell(row=r+3, column=4+len(student_averages[student]['Details'])).value = student_averages[student]['Overall Average']
	
	
	
	for c,grade in enumerate(student_averages[student]['Maths']['Grade Levels']):
		if grade!='Details':
			sheet4.cell(row=2, column=c+len(student_averages[student]['Details'])+1).value = grade
			sheet4.cell(row=r+3, column=c+len(student_averages[student]['Details'])+1).value = student_averages[student]['Maths']['Grade Levels'][grade]
	for c,grade in enumerate(student_averages[student]['Language']['Grade Levels']):
		if grade!='Details':
			sheet4.cell(row=2, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Grade Levels'])).value = grade
			sheet4.cell(row=r+3, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Grade Levels'])).value = student_averages[student]['Language']['Grade Levels'][grade]
	
	for c,grade in enumerate(student_averages[student]['Maths']['Cognitive Domain']):
		if grade!='Details':
			sheet5.cell(row=2, column=c+len(student_averages[student]['Details'])+1).value = grade
			sheet5.cell(row=r+3, column=c+len(student_averages[student]['Details'])+1).value = student_averages[student]['Maths']['Cognitive Domain'][grade]
	for c,grade in enumerate(student_averages[student]['Language']['Cognitive Domain']):
		if grade!='Details':
			sheet5.cell(row=2, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Cognitive Domain'])).value = grade
			sheet5.cell(row=r+3, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Cognitive Domain'])).value = student_averages[student]['Language']['Cognitive Domain'][grade]
	
	for c,grade in enumerate(student_averages[student]['Maths']['Content Domain']):
		if grade!='Details':
			sheet6.cell(row=2, column=c+len(student_averages[student]['Details'])+1).value = grade
			sheet6.cell(row=r+3, column=c+len(student_averages[student]['Details'])+1).value = student_averages[student]['Maths']['Content Domain'][grade]
	for c,grade in enumerate(student_averages[student]['Language']['Content Domain']):
		if grade!='Details':
			sheet6.cell(row=2, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Content Domain'])).value = grade
			sheet6.cell(row=r+3, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Content Domain'])).value = student_averages[student]['Language']['Content Domain'][grade]		
	for c,detail in enumerate(student_averages[student]['Details']):
		if detail != 'Scores':
			sheet3.cell(row=2, column=c+1).value = detail
			sheet3.cell(row=r+3,column = c+1).value = student_averages[student]['Details'][detail]
			sheet4.cell(row=2, column=c+1).value = detail
			sheet4.cell(row=r+3,column = c+1).value = student_averages[student]['Details'][detail]
			sheet5.cell(row=2, column=c+1).value = detail
			sheet5.cell(row=r+3,column = c+1).value = student_averages[student]['Details'][detail]
			sheet6.cell(row=2, column=c+1).value = detail
			sheet6.cell(row=r+3,column = c+1).value = student_averages[student]['Details'][detail]
		
sheet3.cell(row=2, column=1).value= 'Student'
sheet4.cell(row=2, column=1).value= 'Student'
sheet5.cell(row=2, column=1).value= 'Student'
sheet6.cell(row=2, column=1).value= 'Student'

new_book.save('Testing3.xlsx')